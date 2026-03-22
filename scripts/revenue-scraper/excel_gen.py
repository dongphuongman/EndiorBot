"""
Generate Excel (.xlsx) from revenue row data.
"""
from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# Columns to highlight / format as currency (VND)
CURRENCY_KEYS = {
    "doanh_thu", "revenue", "amount", "total", "net", "gross",
    "tong", "doanh thu", "so_tien", "gia_tri",
}


def generate(rows: list[dict[str, Any]], output_path: str, period: str, from_date: date, to_date: date) -> str:
    """
    Generate Excel file from rows data.
    Returns the output path.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    if not rows:
        raise ValueError("No data to export")

    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Detail"

    # ── Title row ──
    title = f"Báo Cáo Doanh Thu — {period.upper()} ({from_date} → {to_date})"
    ws.merge_cells("A1:Z1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="2E4057")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Headers row ──
    headers = list(rows[0].keys())
    HEADER_FILL = PatternFill("solid", fgColor="4A90D9")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # ── Data rows ──
    EVEN_FILL = PatternFill("solid", fgColor="EBF5FB")
    NUMBER_FORMAT = '#,##0'

    for row_idx, row in enumerate(rows, start=3):
        for col_idx, header in enumerate(headers, start=1):
            val = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx)

            # Try to convert numeric strings
            if isinstance(val, str):
                clean = val.replace(",", "").replace(".", "").strip()
                if clean.lstrip("-").isdigit():
                    val = int(clean)
                else:
                    try:
                        val = float(val.replace(",", ""))
                    except (ValueError, AttributeError):
                        pass

            cell.value = val

            # Currency formatting
            if isinstance(val, (int, float)) and header.lower() in CURRENCY_KEYS:
                cell.number_format = NUMBER_FORMAT

            # Zebra striping
            if row_idx % 2 == 0:
                cell.fill = EVEN_FILL

    # ── Auto-fit column widths ──
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(header)),
            *[len(str(rows[r].get(header, ""))) for r in range(min(len(rows), 50))]
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # ── Summary row ──
    summary_row = len(rows) + 3
    ws.cell(row=summary_row, column=1, value=f"Tổng: {len(rows)} dòng").font = Font(italic=True, color="666666")

    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("[excel] Saved %d rows to %s", len(rows), output_path)
    return output_path
